#!/usr/bin/env python3
"""
Kairos Cup 2025 — Build Script
Calcola classifica/cannonieri/portieri in Python dai fogli PARTITE e GIOCATORI.
NON dipende dalle formule Excel (che openpyxl non esegue).

Uso: python build.py
Dipendenze: pip install openpyxl
"""

import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("❌  Installa openpyxl:  pip install openpyxl")

ROOT     = Path(__file__).parent
EXCEL    = ROOT / "KairosCup_Dati.xlsx"
TEMPLATE = ROOT / "index_template.html"
OUT_DIR  = ROOT / "docs"
OUT_HTML = OUT_DIR / "index.html"

ALL_TEAMS = [
    "Torino 1","Torino 2","Mestre 1","Venezia/Mestre",
    "Merano","Pordenone","San Giorgio di Nogaro",
    "Udine","Trieste 1","Trieste 2"
]

def safe_int(val, default=0):
    if val is None or val == "": return default
    try: return int(float(str(val)))
    except: return default

def safe_str(val):
    return str(val).strip() if val is not None else ""

def calc_classifica(ws):
    stats = {t:{"G":0,"V":0,"N":0,"P":0,"GF":0,"GS":0} for t in ALL_TEAMS}
    for row in ws.iter_rows(min_row=4, max_row=48, values_only=True):
        ta = safe_str(row[2]); tb = safe_str(row[6])
        try:
            ga = int(float(str(row[3]))); gb = int(float(str(row[5])))
        except: continue
        if ta not in stats or tb not in stats: continue
        stats[ta]["G"]+=1; stats[tb]["G"]+=1
        stats[ta]["GF"]+=ga; stats[ta]["GS"]+=gb
        stats[tb]["GF"]+=gb; stats[tb]["GS"]+=ga
        if   ga>gb: stats[ta]["V"]+=1; stats[tb]["P"]+=1
        elif ga<gb: stats[tb]["V"]+=1; stats[ta]["P"]+=1
        else:       stats[ta]["N"]+=1; stats[tb]["N"]+=1
    rows = []
    for team,s in stats.items():
        pt=s["V"]*3+s["N"]; dr=s["GF"]-s["GS"]
        rows.append({**s,"squadra":team,"Pt":pt,"Dr":dr})
    rows.sort(key=lambda t:(-t["Pt"],-t["Dr"],-t["GF"]))
    for i,r in enumerate(rows): r["pos"]=i+1
    return rows

def calc_giocatori(ws):
    players=[]
    for row in ws.iter_rows(min_row=3, max_row=202, values_only=True):
        nome=safe_str(row[2])
        if not nome: continue
        players.append({
            "nome":nome,"squadra":safe_str(row[3]),
            "gol":safe_int(row[4]),"assist":safe_int(row[5]),
            "voto":safe_int(row[6]),"presenze":safe_int(row[7]),
            "ruolo":safe_str(row[8]),
        })
    cannonieri=sorted([p for p in players if p["gol"]>0],key=lambda p:(-p["gol"],-p["assist"]))[:10]
    portieri=sorted([p for p in players if p["ruolo"].lower()=="portiere" and p["voto"]>0],key=lambda p:-p["voto"])[:5]
    return (
        [{"nome":p["nome"],"squadra":p["squadra"],"gol":p["gol"],"assist":p["assist"],"voto":p["voto"]} for p in cannonieri],
        [{"nome":p["nome"],"squadra":p["squadra"],"voto":p["voto"]} for p in portieri]
    )

def calc_finali(ws):
    labels=["Finale 1°/2°","Finale 3°/4°","Finale 5°/6°","Finale 7°/8°","Finale 9°/10°"]
    fin_rows=[]
    for row in ws.iter_rows(min_row=49, max_row=62, values_only=True):
        if safe_str(row[1])=="FIN":
            fin_rows.append(row)
    finali=[]
    for i,row in enumerate(fin_rows[:5]):
        ta=safe_str(row[2]); tb=safe_str(row[6])
        try: score=f"{int(float(str(row[3])))}-{int(float(str(row[5])))}"
        except: score=None
        finali.append({
            "label":labels[i] if i<len(labels) else f"Finale {i+1}",
            "teamA":ta if ta and "TBD" not in ta else "TBD",
            "teamB":tb if tb and "TBD" not in tb else "TBD",
            "score":score,
        })
    return finali

def build():
    print(f"📂  Leggo  {EXCEL.name}  …")
    if not EXCEL.exists():    sys.exit(f"❌  File non trovato: {EXCEL}")
    if not TEMPLATE.exists(): sys.exit(f"❌  Template non trovato: {TEMPLATE}")
    try:
        wb=openpyxl.load_workbook(EXCEL, data_only=False)
    except Exception as e:
        sys.exit(f"❌  Impossibile aprire Excel: {e}")
    for s in ("PARTITE","GIOCATORI"):
        if s not in wb.sheetnames: sys.exit(f"❌  Foglio '{s}' non trovato.")

    classifica=calc_classifica(wb["PARTITE"])
    cannonieri,portieri=calc_giocatori(wb["GIOCATORI"])
    finali=calc_finali(wb["PARTITE"])

    data={"classifica":classifica,"cannonieri":cannonieri,"portieri":portieri,"finali":finali}
    print(f"✅  Classifica: {len(classifica)} squadre")
    print(f"✅  Cannonieri: {len(cannonieri)} giocatori")
    print(f"✅  Portieri:   {len(portieri)} portieri")
    print(f"✅  Finali:     {len(finali)} partite")

    template=TEMPLATE.read_text(encoding="utf-8")
    output=template.replace("__KAIROS_DATA__", json.dumps(data, ensure_ascii=False))
    if "__KAIROS_DATA__" in output:
        sys.exit("❌  Placeholder __KAIROS_DATA__ non trovato in index_template.html")

    OUT_DIR.mkdir(exist_ok=True)
    OUT_HTML.write_text(output, encoding="utf-8")
    print(f"\n🚀  Sito generato in:  {OUT_HTML.relative_to(ROOT)}")

if __name__=="__main__":
    build()
